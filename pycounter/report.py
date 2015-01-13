"""COUNTER journal and book reports and associated functions"""

from __future__ import absolute_import

import logging
import re
import warnings
import datetime
import calendar

import pyisbn
import six

from pycounter import _csvhelper

METRICS = {u"JR1": u"FT Article Requests",
           u"BR1": u"Book Title Requests",
           u"BR2": u"Book Section Requests",
           }


class UnknownReportTypeError(Exception):
    """We can't parse this kind of report yet."""
    pass


class CounterReport(object):
    """
    a COUNTER usage statistics report.

    Iterate over the report object to get its rows (each of which is a
    :class:`CounterBook <CounterBook>` or :class:`CounterPublication
    <CounterPublication>` instance.

    :param metric: metric being tracked by this report. For database
        reports (which have multiple metrics per report, and which aren't
        implemented yet), this should be set to `None`.


    Other attributes aren't currently set at creation time, but rather
    are set by the parser after creating the object. (This should
    change in the future.)

    """
    def __init__(self, metric=None):
        self.pubs = []
        self._year = None
        self.report_type = None
        self.report_version = 0
        self.metric = metric
        self.customer = None
        self.institutional_identifier = None
        self.period = (None, None)
        self.date_run = None

    def __str__(self):
        return "CounterReport %s version %s for %s" % (self.report_type,
                                                       self.report_version,
                                                       self.year)

    def __iter__(self):
        return iter(self.pubs)

    @property
    def year(self):
        """year covered by report.

        Raises AttributeError if report is for multiple years
        or if it doesn't begin in January.

        This attribute is deprecated.
        """
        if (self.period[0].month != 1 or
                self.period[0].year != self.period[1].year):
            raise AttributeError("no year attribute for multiyear reports")
        warnings.warn("year attribute is deprecated", DeprecationWarning)
        if self._year is None:
            return self.period[0].year
        else:
            return self._year

    @year.setter
    def year(self, value):
        self._year = value


class CounterEresource(six.Iterator):
    """
    base class for COUNTER statistics lines

    Iterating returns (first_day_of_month, metric, usage) tuples.
    """

    def __init__(self, line=None, period=None, metric=None):
        self.period = period
        self.metric = metric
        """period covered by this report"""
        if line is not None:
            self.title = line[0]
            """title of eResource"""
            self.publisher = line[1]
            """publisher of eResource"""
            self.platform = line[2]
            """platform hosting eResource"""
            self._monthdata = [format_stat(x) for x in line[5:]]
            while len(self._monthdata) < 12:
                self._monthdata.append(None)
            logging.debug("monthdata: %s", self._monthdata)

    @property
    def monthdata(self):
        """
        Return month data when applicable.

        Deprecated. Raises AttributeError for multi-year reports
        """
        if (self.period[0].month != 1 or
                self.period[0].year != self.period[1].year):
            raise AttributeError("no monthdata for multiyear reports")
        warnings.warn("monthdata is deprecated", DeprecationWarning)
        return self._monthdata

    def __iter__(self):
        currmonth = self.period[0]
        md = iter(self._monthdata)
        while currmonth < self.period[1]:
            currusage = next(md)
            yield (currmonth, self.metric, currusage)
            currmonth = _next_month(currmonth)


class CounterPublication(CounterEresource):
    """
    statistics for a single electronic journal.

    :param line: a list containing the usage data for this line, in
        COUNTER 3 layout. (This is an ugly hack that should be fixed
        very soon)

    :param metric: the metric tracked by this statistics line.
        (Should probably always be "FT Article Requests" for
        CounterPublication objects, as long as only JR1 is supported.)

    """
    def __init__(self, line=None, period=None, metric=METRICS[u"JR1"]):
        super(CounterPublication, self).__init__(line, period, metric)
        if line is not None:
            self.issn = line[3].strip()
            """eJournal's print ISSN"""
            self.eissn = line[4].strip()
            """eJournal's eISSN"""
            self.isbn = None

    def __str__(self):
        return """<CounterPublication %s, publisher %s,
        platform %s>""" % (self.title, self.publisher, self.platform)


class CounterBook(CounterEresource):
    """
    statistics for a single electronic book.

    :param line: a list containing the usage data for this line, in
        COUNTER 3 layout. (This is an ugly hack that should be fixed
        very soon)

    """
    def __init__(self, line=None, period=None, metric=None):
        super(CounterBook, self).__init__(line, period, metric)
        if line is not None:
            self.isbn = line[3].strip().replace('-', '')
            """eBook's ISBN"""
            if len(self.isbn) == 10:
                self.isbn = pyisbn.convert(self.isbn)
            self.issn = line[4].strip()
            """eBook's ISSN (if any)"""
            self.eissn = None

    def __str__(self):
        return """<CounterBook %s (ISBN: %s), publisher %s,
        platform %s>""" % (self.title, self.isbn, self.publisher,
                           self.platform)


def format_stat(stat):
    stat = stat.replace(',', '')
    try:
        return int(stat)
    except ValueError:
        return None


def parse(filename):
    """Parse a COUNTER file, first attempting to determine type

    Returns a :class:`CounterReport <CounterReport>` object.

    :param filename: path to COUNTER report to load and parse.
    """
    if filename.endswith('.tsv'):
        # Horrible filename-based hack; in future examine contents of file here
        return parse_separated(filename, '\t')
    if filename.endswith('.xlsx'):
        return parse_xlsx(filename)
    # fallback to old assume-csv behavior
    return parse_separated(filename, ',')


def parse_xlsx(filename):
    """Parse a COUNTER file in Excel format.

    Invoked automatically by ``parse``.

    :param filename: path to XLSX-format COUNTER report file.

    """
    from openpyxl import load_workbook
    workbook = load_workbook(filename=filename)
    worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
    row_it = worksheet.iter_rows()
    split_row_list = ([cell.value if cell.value is not None else ""
                       for cell in row] for row in row_it)

    return parse_generic(split_row_list)


def parse_separated(filename, delimiter):
    """Open COUNTER CSV/TSV report with given filename and delimiter
    and parse into a CounterReport object

    Invoked automatically by ``parse``.

    :param filename: path to delimited COUNTER report file.
    :param delimiter: character (such as ',' or '\t') used as the
        delimiter for this file

    """
    with _csvhelper.UnicodeReader(filename,
                                  delimiter=delimiter) as report_reader:
        return parse_generic(report_reader)


def parse_generic(report_reader):
    """Takes an iterator of COUNTER report rows and
    returns a CounterReport object

    :param report_reader: a iterable object that yields lists COUNTER
        data formatted as tablular lists

    """
    report = CounterReport()

    line1 = six.next(report_reader)

    rt_match = re.match(
        r'.*(Journal|Book|Database) Report (\d) ?\(R(\d)\)',
        line1[0])
    if rt_match:
        report.report_type = (rt_match.group(1)[0].capitalize() + 'R' +
                              rt_match.group(2))
        report.report_version = int(rt_match.group(3))

    report.metric = METRICS.get(report.report_type)

    custline = six.next(report_reader)
    report.customer = custline[0]

    if report.report_version == 4:
        inst_id_line = six.next(report_reader)
        if inst_id_line:
            report.institutional_identifier = inst_id_line[0]

        six.next(report_reader)

        covered_line = six.next(report_reader)
        report.period = _convert_covered(covered_line[0])

    six.next(report_reader)

    date_run_line = six.next(report_reader)
    report.date_run = _convert_date_run(date_run_line[0])

    header = six.next(report_reader)
    first_date_col = 10 if report.report_version == 4 else 5
    if report.report_type in ('BR1', 'BR2') and report.report_version == 4:
        first_date_col = 8

    year = int(header[first_date_col].split('-')[1])
    if year < 100:
        year += 2000

    report.year = year

    if report.report_version == 4:
        countable_header = header[0:8]
        for col in header[8:]:
            if col:
                countable_header.append(col)
        last_col = len(countable_header)
    else:
        for last_col, v in enumerate(header):
            if 'YTD' in v:
                break
        start_date = datetime.date(year, 1, 1)
        end_date = _last_day(_convert_date_column(header[last_col - 1]))
        report.period = (start_date, end_date)

    six.next(report_reader)
    for line in report_reader:
        if not line:
            continue
        if report.report_version == 4:
            if report.report_type == 'JR1':
                line = line[0:3] + line[5:7] + line[10:last_col]
            elif report.report_type in ('BR1', 'BR2'):
                line = line[0:3] + line[5:7] + line[8:last_col]
        else:
            line = line[0:last_col]
        logging.debug(line)
        if report.report_type:
            if report.report_type.startswith('JR'):
                report.pubs.append(CounterPublication(line,
                                                      report.period,
                                                      report.metric))
            elif report.report_type.startswith('BR'):
                report.pubs.append(CounterBook(line,
                                               report.period,
                                               report.metric))
            else:
                raise UnknownReportTypeError(report.report_type)

    return report


def parse_csv(filename):
    """Parse CSV files; deprecated."""
    warnings.warn(".parse_csv is deprecated; use parse_separated",
                  DeprecationWarning)
    return parse_separated(filename, ",")


def parse_tsv(filename):
    """Parse tab-separated files; deprecated."""
    warnings.warn(".parse_tsv is deprecated; use parse_separated",
                  DeprecationWarning)
    return parse_separated(filename, "\t")


def _convert_covered(datestring):
    """
    Convert a string of the format 'YYYY-MM-DD to YYYY-MM-DD' to a
    tuple of datetime.date instances.

    :param datestring: the string to convert to a date.

    """
    start_string, end_string = datestring.split(" to ")
    start_date = datetime.datetime.strptime(start_string, "%Y-%m-%d").date()
    end_date = datetime.datetime.strptime(end_string, "%Y-%m-%d").date()

    return (start_date, end_date)


def _convert_date_run(datestring):
    """
    Convert a date of the format 'YYYY-MM-DD' to a datetime.date object

    :param datestring: the string to convert to a date.

    """
    return datetime.datetime.strptime(datestring, "%Y-%m-%d").date()


def _convert_date_column(datestring):
    """
    Convert a month expressed as, e.g., 'Jan-2014' to a datetime.date
    object representing the first day of that month/

    :param datestring: the string to convert to a date.

    """
    return datetime.datetime.strptime(datestring, "%b-%Y").date()


def _last_day(orig_date):
    """
    Return a datetime.date object representing the last day of a
    calendar month, given a datetime.date for any day in that month

    :param orig_date: the date within the month for which we want the
        last day.

    """
    daynum = calendar.monthrange(orig_date.year, orig_date.month)[1]
    return datetime.date(orig_date.year, orig_date.month, daynum)


def _next_month(dateobj):
    """Return a datetime.date for the first day of the next month
    after the given date

    :param orig_date: the date within the month for which we want the
        next month's first day.

    """
    year_delta, prev_month = divmod(dateobj.month, 12)
    return datetime.date(dateobj.year + year_delta, prev_month + 1, 1)
